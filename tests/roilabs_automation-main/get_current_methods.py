import json
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_report_separate_tabs(json_file_path, output_file=None):
    """Create an Excel file with separate tabs for payment and withdrawal methods"""

    # Auto-generate filename using current date if not provided
    if output_file is None:
        today = datetime.today().strftime("%Y_%m_%d")
        output_file = f"current_methods_{today}.xlsx"

    # Load JSON
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: File '{json_file_path}' not found.")
        return
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in file '{json_file_path}'.")
        return

    # Create workbook
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

    # Process each project
    for project in data["projects"]:
        project_name = project["project"]

        countries = []
        all_payment_methods = {}
        all_withdrawal_methods = {}
        has_withdrawal_methods = False

        for country_data in project["countries"]:
            country_name = country_data["country"]
            countries.append(country_name)

            # Payment methods
            payment_methods = []
            if "payment_methods" in country_data and country_data["payment_methods"]:
                payment_methods = [pm['payment_method'] for pm in country_data["payment_methods"]]
            all_payment_methods[country_name] = payment_methods

            # Withdrawal methods
            withdrawal_methods = []
            if "withdraw_methods" in country_data and country_data["withdraw_methods"]:
                withdrawal_methods = [wm['withdraw_method'] for wm in country_data["withdraw_methods"]]
                has_withdrawal_methods = True
            all_withdrawal_methods[country_name] = withdrawal_methods

        # Create Payments tab
        payment_worksheet = workbook.create_sheet(title=f"{project_name}_Payments")
        create_methods_tab(payment_worksheet, countries, all_payment_methods)
        format_methods_tab(payment_worksheet, countries, "payment")

        # Create Withdrawals tab only if available
        if has_withdrawal_methods:
            withdrawal_worksheet = workbook.create_sheet(title=f"{project_name}_Withdrawals")
            create_methods_tab(withdrawal_worksheet, countries, all_withdrawal_methods)
            format_methods_tab(withdrawal_worksheet, countries, "withdrawal")

    # Save file
    workbook.save(output_file)
    print(f"Excel report generated: {output_file}")


def create_methods_tab(worksheet, countries, methods_data):
    """Create a tab with methods listed under each country"""

    # Headers
    for col, country in enumerate(countries, 1):
        worksheet.cell(row=1, column=col, value=country)

    # Methods
    for country in countries:
        col = countries.index(country) + 1
        row = 2

        for method in methods_data[country]:
            worksheet.cell(row=row, column=col, value=method)
            row += 1


def format_methods_tab(worksheet, countries, method_type):
    """Apply formatting to the worksheet"""

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Colors
    method_font = Font(color="2E7D32") if method_type == "payment" else Font(color="ED7D31")

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header formatting
    for col in range(1, len(countries) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Method formatting
    for row in range(2, worksheet.max_row + 1):
        for col in range(1, len(countries) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = method_font
            cell.alignment = left_align
            cell.border = thin_border

    # Auto column width
    for col in range(1, len(countries) + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, worksheet.max_row + 1):
            val = worksheet.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)


if __name__ == "__main__":
    json_file_path = 'current_methods.json'
    create_excel_report_separate_tabs(json_file_path)
